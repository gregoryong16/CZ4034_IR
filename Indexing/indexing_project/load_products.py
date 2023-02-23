# from indexing_project.models import Products
import openpyxl
from .models import Products

reader = openpyxl.load_workbook("../static/shopee_products.xlsx")
worksheet = reader.active
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")  
    print('')

for dbframe in worksheet.itertuples():
    obj = Products.objects.create(name=dbframe.name, price = "$2.00", description=dbframe.description,
                                            ratings=dbframe.rating, image_url=dbframe.image_url, shop_id=dbframe.shopid, item_id=dbframe.itemid
                                        )           
    obj.save()

print(Products.objects.filter(name__startswith='SG'))
